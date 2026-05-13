"""
Collecte des settlement prices EEX pour Power FR et Natural Gas TTF.
Deux onglets Excel : "Power FR" et "Natural Gas TTF".
"""

import os
import sys
import time
import requests
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_URL   = "https://api.eex-group.com/pub/market-data/price-ticker"
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "eex_prix.xlsx")

REQUEST_HEADERS = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Origin": "https://www.eex.com",
    "Referer": "https://www.eex.com/",
}

HEADERS = ["Date", "Contrat", "Type", "Prix (€/MWh)", "Source"]

# Onglet par commodité
SHEET_MAP = {
    "Power FR":        ["Power FR Baseload", "Power FR Peakload"],
    "Natural Gas TTF": ["Natural Gas TTF"],
}

# Largeurs de colonnes : Date, Contrat, Type, Prix, Source
COL_WIDTHS = [13, 18, 7, 15, 13]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_THIN       = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ROW_ALT     = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
ROW_NORM    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
REPORTED    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# ---------------------------------------------------------------------------
# Contrats
# ---------------------------------------------------------------------------

def build_contracts():
    today = date.today()
    contracts = []

    # Power FR Baseload ── CAL 2027-2030
    for year in range(2027, 2031):
        contracts.append({
            "label": f"ELEC-CAL-{year}", "commodite": "Power FR Baseload", "type": "CAL",
            "params": {"shortCode": "F7BY", "area": "FR", "product": "Base",
                       "commodity": "POWER", "pricing": "F", "maturity": f"{year}01"},
        })

    # Power FR Baseload ── QTR 2027-2030
    for year in range(2027, 2031):
        for q, month in enumerate(["01", "04", "07", "10"], start=1):
            contracts.append({
                "label": f"ELEC-Q{q}-{year}", "commodite": "Power FR Baseload", "type": "QTR",
                "params": {"shortCode": "F7BQ", "area": "FR", "product": "Base",
                           "commodity": "POWER", "pricing": "F", "maturity": f"{year}{month}"},
            })

    # Power FR Baseload ── MTH : mois courant + 17 mois suivants
    for i in range(18):
        m = today + relativedelta(months=i)
        contracts.append({
            "label": f"ELEC-{m.strftime('%b')}-{m.year}", "commodite": "Power FR Baseload", "type": "MTH",
            "params": {"shortCode": "F7BM", "area": "FR", "product": "Base",
                       "commodity": "POWER", "pricing": "F", "maturity": m.strftime("%Y%m")},
        })

    # Power FR Peakload ── CAL 2027-2030
    for year in range(2027, 2031):
        contracts.append({
            "label": f"PEAK-CAL-{year}", "commodite": "Power FR Peakload", "type": "CAL",
            "params": {"shortCode": "F7PY", "area": "FR", "product": "Peak",
                       "commodity": "POWER", "pricing": "F", "maturity": f"{year}01"},
        })

    # Natural Gas TTF ── CAL 2027-2030
    for year in range(2027, 2031):
        contracts.append({
            "label": f"GAZ-CAL-{year}", "commodite": "Natural Gas TTF", "type": "CAL",
            "params": {"shortCode": "G3BY", "area": "TTF", "product": "Physical",
                       "commodity": "NATGAS", "pricing": "F", "maturity": f"{year}01"},
        })

    # Natural Gas TTF ── QTR 2027-2030
    for year in range(2027, 2031):
        for q, month in enumerate(["01", "04", "07", "10"], start=1):
            contracts.append({
                "label": f"GAZ-Q{q}-{year}", "commodite": "Natural Gas TTF", "type": "QTR",
                "params": {"shortCode": "G3BQ", "area": "TTF", "product": "Physical",
                           "commodity": "NATGAS", "pricing": "F", "maturity": f"{year}{month}"},
            })

    # Natural Gas TTF ── MTH : mois courant + 17 mois suivants
    for i in range(18):
        m = today + relativedelta(months=i)
        contracts.append({
            "label": f"GAZ-{m.strftime('%b')}-{m.year}", "commodite": "Natural Gas TTF", "type": "MTH",
            "params": {"shortCode": "G3BM", "area": "TTF", "product": "Physical",
                       "commodity": "NATGAS", "pricing": "F", "maturity": m.strftime("%Y%m")},
        })

    return contracts


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

def fetch_settlement(params: dict) -> float | None:
    try:
        r = requests.get(BASE_URL, params=params, headers=REQUEST_HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data")
        if not rows:
            return None
        headers = data.get("header", [])
        idx = headers.index("settlPx") if "settlPx" in headers else 1
        price = rows[0][idx]
        if price is None or price == "":
            return None
        return float(price)
    except Exception:
        return None
    finally:
        time.sleep(0.4)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def commodite_to_sheet(commodite: str) -> str:
    for sheet_name, commodites in SHEET_MAP.items():
        if commodite in commodites:
            return sheet_name
    return list(SHEET_MAP.keys())[0]


def _init_sheet(ws):
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def format_sheet(ws):
    """Applique couleurs alternées, bordures, format nombre à toutes les lignes de données."""
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_reported = len(row) > 4 and row[4].value == "Reporté J-1"
        is_alt      = (row_idx % 2 == 0)
        fill        = REPORTED if is_reported else (ROW_ALT if is_alt else ROW_NORM)

        for col_idx, cell in enumerate(row, start=1):
            cell.fill   = fill
            cell.border = BORDER
            if col_idx == 4:  # Prix
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = ws.dimensions


def load_or_create_workbook(path: str):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        # Migration depuis l'ancien format mono-feuille
        for old in ["Prices", "Sheet"]:
            if old in wb.sheetnames:
                _migrate_old_sheet(wb, wb[old])
                del wb[old]
        # Créer les onglets manquants
        for sheet_name in SHEET_MAP:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                _init_sheet(ws)
    else:
        wb = openpyxl.Workbook()
        first = True
        for sheet_name in SHEET_MAP:
            ws = wb.active if first else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            _init_sheet(ws)
            first = False
    return wb


def _migrate_old_sheet(wb, old_ws):
    """Transfère les données de l'ancienne feuille vers les nouveaux onglets."""
    for row in old_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        # Ancien format : Date, Contrat, Commodite, Type, Prix, Source
        if len(row) >= 6:
            date_v, label, commodite, type_v, price, source = row[0], row[1], row[2], row[3], row[4], row[5]
        elif len(row) >= 5:
            date_v, label, type_v, price, source = row[0], row[1], row[2], row[3], row[4]
            commodite = "Power FR Baseload"
        else:
            continue
        sheet_name = commodite_to_sheet(commodite or "")
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _init_sheet(ws)
        wb[sheet_name].append([date_v, label, type_v, price, source or "Settlement"])


def already_has_today(wb, today_str: str) -> bool:
    for sheet_name in SHEET_MAP:
        if sheet_name not in wb.sheetnames:
            continue
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if row and str(row[0]) == today_str:
                return True
    return False


def get_last_known_price(ws, label: str) -> float | None:
    last_price = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1] == label and row[3] is not None:
            last_price = row[3]
    return last_price


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    today     = date.today()
    today_str = today.strftime("%Y-%m-%d")

    print(f"\n=== Collecte EEX — {today_str} ===")

    contracts = build_contracts()

    os.makedirs(os.path.dirname(os.path.abspath(EXCEL_PATH)), exist_ok=True)
    wb = load_or_create_workbook(EXCEL_PATH)

    if already_has_today(wb, today_str):
        print("  Données du jour déjà présentes — rien à faire.")
        sys.exit(0)

    added = reported = ignored = 0

    for c in contracts:
        price      = fetch_settlement(c["params"])
        sheet_name = commodite_to_sheet(c["commodite"])
        ws         = wb[sheet_name]

        if price is not None:
            ws.append([today_str, c["label"], c["type"], price, "Settlement"])
            print(f"  {c['label']}: {price} €/MWh")
            added += 1
        else:
            last = get_last_known_price(ws, c["label"])
            if last is not None:
                ws.append([today_str, c["label"], c["type"], last, "Reporté J-1"])
                print(f"  {c['label']}: {last} €/MWh ⚠️ reporté")
                reported += 1
            else:
                print(f"  {c['label']}: ignoré (pas de prix)")
                ignored += 1

    for sheet_name in SHEET_MAP:
        if sheet_name in wb.sheetnames:
            format_sheet(wb[sheet_name])

    wb.save(EXCEL_PATH)
    print(f"\n✓ Terminé — {added} lignes ajoutées ({reported} reportées, {ignored} ignorées)")


if __name__ == "__main__":
    main()
